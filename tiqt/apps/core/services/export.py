"""Exportação de relatórios para Excel e PDF.

Ambos consomem a mesma estrutura `secoes` — uma lista de dicts com titulo,
colunas e linhas. Assim o Excel e o PDF nunca divergem no conteúdo: se um mostra
um número, o outro mostra o mesmo.

Excel: openpyxl (já estava no requirements, com um import morto em views.py).
PDF: xhtml2pdf — pip puro, sem dependência de sistema. WeasyPrint gera PDF mais
bonito mas exige GTK/Pango instalado no SO, o que quebra no Windows e vira
dependência de deploy no Linux.
"""

import io
from datetime import date, datetime

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone


def _fmt(v):
    if v is None:
        return '—'
    if isinstance(v, float):
        return f'{v:.1f}'
    return v


def minutos_legivel(m):
    """Minutos úteis -> algo que um humano lê sem fazer conta.

    Um SLA em 'minutos úteis' é preciso e ilegível: 6918min não diz nada, '11,8
    dias úteis' diz. O dia útil aqui é o expediente cadastrado, não 24h.
    """
    from .sla import carregar_calendario, minutos_por_dia_util

    if m is None:
        return '—'
    if m < 60:
        return f'{int(m)}min'

    minutos_dia = minutos_por_dia_util(carregar_calendario())
    if not minutos_dia or m < minutos_dia:
        return f'{m / 60:.1f}h'
    return f'{m / minutos_dia:.1f}d úteis'


def to_excel(secoes, nome='relatorio'):
    """Uma aba por seção."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    cab_fill = PatternFill('solid', fgColor='1a1a19')
    cab_font = Font(bold=True, color='FFFFFF')

    for sec in secoes:
        # Nome de aba no Excel: 31 chars, sem : \ / ? * [ ]
        titulo = sec['titulo'][:31]
        for c in ':\\/?*[]':
            titulo = titulo.replace(c, '-')
        ws = wb.create_sheet(titulo)

        ws.append(sec['colunas'])
        for cell in ws[1]:
            cell.fill = cab_fill
            cell.font = cab_font
            cell.alignment = Alignment(horizontal='center')

        for linha in sec['linhas']:
            ws.append([_fmt(v) for v in linha])

        # Largura pela maior célula, com teto para não virar aba de 300 colunas.
        for i, _ in enumerate(sec['colunas'], start=1):
            largura = max(
                [len(str(sec['colunas'][i - 1]))]
                + [len(str(_fmt(l[i - 1]))) for l in sec['linhas'] if len(l) >= i]
            )
            ws.column_dimensions[get_column_letter(i)].width = min(largura + 3, 55)

        ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{nome}_{timezone.localdate():%Y-%m-%d}.xlsx"'
    return resp


def to_pdf(secoes, titulo, subtitulo='', avisos=None, nome='relatorio'):
    from xhtml2pdf import pisa

    html = render_to_string('core/relatorio_pdf.html', {
        'titulo': titulo,
        'subtitulo': subtitulo,
        'secoes': secoes,
        'avisos': avisos or [],
        'gerado_em': timezone.localtime(),
    })

    buf = io.BytesIO()
    # encoding explícito: sem isso acentuação vira lixo no PDF.
    erro = pisa.CreatePDF(io.StringIO(html), dest=buf, encoding='utf-8')
    if erro.err:
        return HttpResponse('Falha ao gerar o PDF.', status=500)

    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{nome}_{timezone.localdate():%Y-%m-%d}.pdf"'
    return resp
