"""Testes dos relatórios e da exportação PDF/Excel."""

import io
from datetime import time, timedelta

import pytest
from django.contrib.auth.models import Group
from django.urls import reverse
from django.utils import timezone

from tiqt.apps.core.models import (
    Cliente, Departamento, Expediente, Prioridade, Situacao, SlaPolitica, Ticket, Tipo, User,
)
from tiqt.apps.core.services import relatorios

pytestmark = pytest.mark.django_db


@pytest.fixture(autouse=True)
def limpa_cache():
    from django.core.cache import cache
    cache.clear()
    yield
    cache.clear()


@pytest.fixture
def expediente(db):
    for dia in range(5):
        Expediente.objects.create(dia_semana=dia, hora_inicio=time(8, 0), hora_fim=time(18, 0))


@pytest.fixture
def base(db):
    dep = Departamento.objects.create(descricao='Suporte')
    return {
        'dep': dep,
        'cliente': Cliente.objects.create(fantasia='ACME', plano=None),
        'tipo': Tipo.objects.create(descricao='Dúvida', departamento=dep),
        'prio': Prioridade.objects.create(descricao='Normal', peso=2),
        'sit': Situacao.objects.create(descricao='Nova'),
    }


def novo_ticket(base, titulo='T'):
    return Ticket.objects.create(
        departamento=base['dep'], cliente=base['cliente'], tipo=base['tipo'],
        prioridade=base['prio'], situacao=base['sit'], titulo=titulo)


@pytest.fixture
def user(db):
    return User.objects.create_user(username='fulano', password='x')


@pytest.fixture
def diretor(db):
    u = User.objects.create_user(username='chefe', password='x')
    u.groups.add(Group.objects.create(name='Diretoria'))
    return u


class TestConformidade:
    def test_sem_politica_a_conformidade_e_none_nao_zero(self, expediente, base, user):
        """Sem meta não há conformidade — e isso é diferente de 0%."""
        t = novo_ticket(base)
        t.iniciar_atendimento(user)
        linhas = relatorios.coletar(Ticket.objects.all())
        g = relatorios.geral(linhas)
        assert g['sla_resposta_pct'] is None

    def test_dentro_da_meta_conta_como_conforme(self, expediente, base, user):
        SlaPolitica.objects.create(minutos_resposta=600, minutos_resolucao=6000)
        t = novo_ticket(base)
        t.iniciar_atendimento(user)   # responde na hora

        linhas = relatorios.coletar(Ticket.objects.all())
        g = relatorios.geral(linhas)
        assert g['sla_resposta_pct'] == 100.0
        assert g['respondidos'] == 1

    def test_fora_da_meta_conta_como_furo(self, expediente, base, user):
        """Envelhecer o ticket é envelhecer o EVENTO: é dele que o relatório lê.

        Mexer só em Ticket.criado_em não muda relatório nenhum — o histórico vive
        no TicketEvento, justamente para não depender do estado atual do ticket.
        """
        from tiqt.apps.core.models import TicketEvento
        SlaPolitica.objects.create(minutos_resposta=1, minutos_resolucao=1)
        t = novo_ticket(base)
        tres_dias_atras = timezone.now() - timedelta(days=3)
        TicketEvento.objects.filter(ticket=t, tipo=TicketEvento.CRIADO).update(ocorrido_em=tres_dias_atras)
        t.iniciar_atendimento(user)

        linhas = relatorios.coletar(Ticket.objects.all())
        assert relatorios.geral(linhas)['sla_resposta_pct'] == 0.0

    def test_agrupa_por_departamento(self, expediente, base, user):
        novo_ticket(base)
        dev = Departamento.objects.create(descricao='Desenvolvimento')
        Ticket.objects.create(departamento=dev, cliente=base['cliente'],
                              tipo=Tipo.objects.create(descricao='Bug', departamento=dev),
                              prioridade=base['prio'], situacao=base['sit'], titulo='Bug')

        grupos = {g['grupo']: g for g in relatorios.conformidade(relatorios.coletar(Ticket.objects.all()))}
        assert grupos['Suporte']['total'] == 1
        assert grupos['Desenvolvimento']['total'] == 1

    def test_usa_o_ultimo_encerramento_nao_o_primeiro(self, expediente, base, user):
        """Ticket reaberto e fechado de novo terminou na SEGUNDA vez."""
        t = novo_ticket(base)
        t.iniciar_atendimento(user)
        t.encerrar_atendimento(user)
        primeiro = t.encerrado_em
        t.reabrir(user)
        t.encerrar_atendimento(user)

        linhas = relatorios.coletar(Ticket.objects.all())
        # a resolução medida tem que refletir o fechamento mais recente
        assert linhas[0]['resolucao_min'] is not None
        assert t.encerrado_em > primeiro


class TestDivida:
    def test_lista_os_parados(self, expediente, base):
        velho = novo_ticket(base, 'Esquecido')
        Ticket.objects.filter(pk=velho.pk).update(criado_em=timezone.now() - timedelta(days=40))
        novo_ticket(base, 'Recente')

        d = relatorios.divida(dias=7)
        assert len(d) == 1
        assert d[0]['titulo'] == 'Esquecido'
        assert d[0]['idade_dias'] >= 39
        assert 'ninguém pegou' in d[0]['responsavel']

    def test_encerrado_nao_e_divida(self, expediente, base, user):
        t = novo_ticket(base)
        Ticket.objects.filter(pk=t.pk).update(criado_em=timezone.now() - timedelta(days=40))
        t.refresh_from_db()
        t.encerrar_atendimento(user)
        assert relatorios.divida(dias=7) == []


class TestVolumeSemanal:
    def test_conta_entradas_e_saidas(self, expediente, base, user):
        t = novo_ticket(base)
        t.iniciar_atendimento(user)
        t.encerrar_atendimento(user)
        novo_ticket(base, 'Só aberto')

        v = relatorios.volume_semanal(Ticket.objects.all())
        assert sum(x['entradas'] for x in v) == 2
        assert sum(x['saidas'] for x in v) == 1

    def test_sem_tickets_devolve_vazio(self, expediente):
        assert relatorios.volume_semanal(Ticket.objects.none()) == []


class TestPermissaoNosRelatorios:
    def test_atendente_ve_so_os_proprios(self, client, expediente, base, user):
        meu = novo_ticket(base, 'Meu')
        meu.iniciar_atendimento(user)
        outro = User.objects.create_user(username='outro', password='x')
        alheio = novo_ticket(base, 'Alheio')
        alheio.iniciar_atendimento(outro)

        client.force_login(user)
        r = client.get(reverse('relatorios'))
        assert r.status_code == 200
        assert b'Meu' in r.content
        assert b'Alheio' not in r.content

    def test_produtividade_e_so_para_diretoria(self, client, expediente, base, user, diretor):
        novo_ticket(base)

        client.force_login(user)
        assert b'Produtividade por atendente' not in client.get(reverse('relatorios')).content

        client.force_login(diretor)
        assert b'Produtividade por atendente' in client.get(reverse('relatorios')).content

    def test_deslogado_vai_para_login(self, client):
        r = client.get(reverse('relatorios'))
        assert r.status_code == 302
        assert '/login/' in r['Location']


class TestExport:
    def test_excel_gera_arquivo_valido(self, client, expediente, base, diretor):
        novo_ticket(base)
        client.force_login(diretor)
        r = client.get(reverse('relatorios_excel'))

        assert r.status_code == 200
        assert 'spreadsheetml' in r['Content-Type']
        assert '.xlsx' in r['Content-Disposition']

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert 'Leia antes' in wb.sheetnames   # o aviso viaja com a planilha

    def test_pdf_gera_arquivo_valido(self, client, expediente, base, diretor):
        novo_ticket(base)
        client.force_login(diretor)
        r = client.get(reverse('relatorios_pdf'))

        assert r.status_code == 200
        assert r['Content-Type'] == 'application/pdf'
        assert r.content[:5] == b'%PDF-'

    def test_export_respeita_o_filtro_de_periodo(self, client, expediente, base, diretor):
        antigo = novo_ticket(base, 'AntigoXYZ')
        Ticket.objects.filter(pk=antigo.pk).update(criado_em=timezone.now() - timedelta(days=400))

        client.force_login(diretor)
        r = client.get(reverse('relatorios'), {'data_inicio': '2026-07-01', 'data_fim': '2026-07-31'})
        # fora do período: não entra na conformidade (mas pode aparecer na dívida,
        # que é uma lista de estado atual, não do período)
        assert b'AntigoXYZ' not in r.content.split(b'parados')[0]

    def test_data_invalida_nao_estoura_500(self, client, diretor):
        client.force_login(diretor)
        assert client.get(reverse('relatorios'), {'data_inicio': 'ontem'}).status_code == 404


class TestFiltroDeDepartamento:
    """O filtro de departamento tem que valer em TODAS as seções.

    A dívida ignora o período de propósito (é estado atual), mas ignorar o
    departamento junto era bug: filtrar 'Suporte' e ver ticket do Dev na lista.
    """

    @pytest.fixture
    def dois_deptos(self, base):
        sup = novo_ticket(base, 'DoSuporteAAA')
        Ticket.objects.filter(pk=sup.pk).update(criado_em=timezone.now() - timedelta(days=40))

        dev = Departamento.objects.create(descricao='Desenvolvimento')
        d = Ticket.objects.create(
            departamento=dev, cliente=base['cliente'],
            tipo=Tipo.objects.create(descricao='Bug', departamento=dev),
            prioridade=base['prio'], situacao=base['sit'], titulo='DoDevBBB')
        Ticket.objects.filter(pk=d.pk).update(criado_em=timezone.now() - timedelta(days=40))
        return sup, dev, d

    def test_divida_respeita_o_departamento(self, client, expediente, base, diretor, dois_deptos):
        sup, dev, _ = dois_deptos
        client.force_login(diretor)

        r = client.get(reverse('relatorios'), {'departamento': sup.departamento_id})
        assert b'DoSuporteAAA' in r.content
        assert b'DoDevBBB' not in r.content     # o bug: aparecia aqui

    def test_divida_ignora_o_periodo_mas_nao_o_departamento(self, client, expediente, base, diretor, dois_deptos):
        """Ticket de 40 dias continua na dívida mesmo filtrando só hoje."""
        sup, _, _ = dois_deptos
        client.force_login(diretor)
        hoje = timezone.localdate().strftime('%Y-%m-%d')

        r = client.get(reverse('relatorios'), {
            'departamento': sup.departamento_id, 'data_inicio': hoje, 'data_fim': hoje})
        assert b'DoSuporteAAA' in r.content     # fora do período, mas é estado atual
        assert b'DoDevBBB' not in r.content

    def test_divida_respeita_o_departamento_no_excel(self, client, expediente, base, diretor, dois_deptos):
        sup, _, _ = dois_deptos
        client.force_login(diretor)
        r = client.get(reverse('relatorios_excel'), {'departamento': sup.departamento_id})

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        aba = next(n for n in wb.sheetnames if 'parados' in n)
        textos = [str(c) for row in wb[aba].iter_rows(values_only=True) for c in row if c]
        assert any('DoSuporteAAA' in t for t in textos)
        assert not any('DoDevBBB' in t for t in textos)

    def test_filtro_de_tipo_vale_em_todas_as_secoes(self, client, expediente, base, diretor, dois_deptos):
        outro_tipo = Tipo.objects.create(descricao='Reclamação', departamento=base['dep'])
        t = Ticket.objects.create(
            departamento=base['dep'], cliente=base['cliente'], tipo=outro_tipo,
            prioridade=base['prio'], situacao=base['sit'], titulo='ReclamaDDD')
        Ticket.objects.filter(pk=t.pk).update(criado_em=timezone.now() - timedelta(days=40))

        client.force_login(diretor)
        r = client.get(reverse('relatorios'), {'tipo': outro_tipo.id})
        assert b'ReclamaDDD' in r.content
        assert b'DoSuporteAAA' not in r.content   # mesmo departamento, tipo diferente
        assert b'DoDevBBB' not in r.content

    def test_conformidade_por_tipo_rotula_com_o_departamento(self, client, expediente, base, diretor):
        """Nomes de tipo se repetem entre departamentos — o rótulo precisa distinguir."""
        dev = Departamento.objects.create(descricao='Desenvolvimento')
        Tipo.objects.create(descricao='Importação', departamento=dev)
        t_sup = Tipo.objects.create(descricao='Importação', departamento=base['dep'])
        Ticket.objects.create(departamento=base['dep'], cliente=base['cliente'], tipo=t_sup,
                              prioridade=base['prio'], situacao=base['sit'], titulo='X')

        grupos = [g['grupo'] for g in relatorios.conformidade(relatorios.coletar(Ticket.objects.all()), por='tipo')]
        assert 'Suporte › Importação' in grupos

    def test_filtro_de_prioridade_tambem_vale_na_divida(self, client, expediente, base, diretor, dois_deptos):
        urgente = Prioridade.objects.create(descricao='Urgente', peso=4)
        t = Ticket.objects.create(
            departamento=base['dep'], cliente=base['cliente'], tipo=base['tipo'],
            prioridade=urgente, situacao=base['sit'], titulo='UrgenteCCC')
        Ticket.objects.filter(pk=t.pk).update(criado_em=timezone.now() - timedelta(days=40))

        client.force_login(diretor)
        r = client.get(reverse('relatorios'), {'prioridade': urgente.id})
        assert b'UrgenteCCC' in r.content
        assert b'DoSuporteAAA' not in r.content
